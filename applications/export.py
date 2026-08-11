"""Экспорт отчётов в Excel (openpyxl) и PDF (reportlab)."""
import datetime
import os

from django.http import HttpResponse


# ---------------------------------------------------------------------------
# Поиск шрифта с поддержкой кириллицы для PDF
# ---------------------------------------------------------------------------
CANDIDATE_FONTS = [
    r'C:\Windows\Fonts\arial.ttf',
    r'C:\Windows\Fonts\DejaVuSans.ttf',
    r'/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    r'/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
]


def _find_font():
    for path in CANDIDATE_FONTS:
        if os.path.exists(path):
            return path
    return None


def excel_response(title, headers, rows, filename):
    """Отдаёт xlsx-файл. rows — список списков (могут включать заголовок итого)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = (title or 'Отчёт')[:31]

    ws.append([title])
    ws['A1'].font = Font(bold=True, size=14)
    ws.append([f'Сформирован: {datetime.datetime.now():%d.%m.%Y %H:%M}'])
    ws.append([])

    if headers:
        ws.append(headers)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='0EA5E9', end_color='0EA5E9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')

    for row in rows:
        ws.append([('' if v is None else v) for v in row])

    # Автоширина колонок
    for col in range(1, (len(headers) if headers else 1) + 1):
        letter = get_column_letter(col)
        max_len = 0
        for cell in ws[letter][4:]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def pdf_response(title, headers, rows, filename, landscape=True):
    """Отдаёт PDF с таблицей. Шрифт с кириллицей — arial/DejaVu."""
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = 'Helvetica'
    font_path = _find_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont('EPFont', font_path))
            font_name = 'EPFont'
        except Exception:
            font_name = 'Helvetica'

    page_size = rl_landscape(A4) if landscape else A4
    buf = HttpResponse(content_type='application/pdf')
    buf['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(buf, pagesize=page_size, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    styles['Title'].fontName = font_name
    styles['Title'].fontSize = 16
    styles['Title'].spaceAfter = 2
    styles['Normal'].fontName = font_name

    story = []
    story.append(Paragraph(title, styles['Title']))
    story.append(Paragraph(f'Сформирован: {datetime.datetime.now():%d.%m.%Y %H:%M}', styles['Normal']))
    story.append(Spacer(1, 8))

    data = []
    if headers:
        data.append(headers)
    for row in rows:
        data.append([('' if v is None else str(v)) for v in row])

    if data:
        table = Table(data, repeatRows=1 if headers else 0)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), font_name),
            ('FONTNAME', (0, 1), (-1, -1), font_name),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0EA5E9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(table)

    doc.build(story)
    return buf
