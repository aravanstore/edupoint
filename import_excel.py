# -*- coding: utf-8 -*-
"""Импорт реальных групп и учеников из Excel «Все группы и студенты».

Запуск:
    python import_excel.py --dry-run     # только разбор и вывод, без записи в БД
    python import_excel.py --commit      # реальный импорт

Формат файла: 4 листа = 4 аудитории. В каждом листе — сетка времени
(09-00/14-00, 10-00/15-00, ...), под каждым временем строка
"Преподаватель + предмет/книга", затем таблица ФИО/телефон/день оплаты.
"""
import os
import re
import sys
import random
import string
import argparse

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = r"D:\Desktop\б\Все группы и студенты (1).xlsx"

env_path = os.path.join(BASE, 'deploy.env')
if os.path.exists(env_path):
    with open(env_path, encoding='utf-8-sig') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                k, v = line.split('=', 1)
                os.environ.setdefault(k.strip(), v.strip())

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'edupoint.settings')
import django
django.setup()

import openpyxl
from datetime import time as dtime

TIME_RE = re.compile(r'^\d{1,2}-\d{2}$')
DIGIT_RE = re.compile(r'\d+')
CEFR_RE = re.compile(r'\b([ABАВ][12])\b', re.I)
CEFR_NORMALIZE = {'А': 'A', 'В': 'B'}  # кириллица -> латиница

TEACHER_DEFAULT_CATEGORY = {
    'нускайым': 'korean',
    'асема': 'english',   # переопределяется на chinese, если явно упомянут китайский
    'элнура': 'english',
    'айгерим': 'german',
    'айдай': 'english',
}

CATEGORY_META = {
    'korean': ('Корейский', '🇰🇷', 'korean'),
    'english': ('Английский', '🇬🇧', 'english'),
    'german': ('Немецкий', '🇩🇪', 'german'),
    'chinese': ('Китайский', '🇨🇳', 'chinese'),
}


def parse_time(marker):
    h, m = marker.split('-')
    return dtime(int(h), int(m))


def detect_category(subject_text, teacher_key):
    t = subject_text.lower()
    if 'немецк' in t:
        return 'german'
    if 'китайск' in t:
        return 'chinese'
    if 'китеп' in t or 'книга' in t:
        return 'korean'
    if CEFR_RE.search(subject_text):
        return 'english'
    return TEACHER_DEFAULT_CATEGORY.get(teacher_key, 'english')


def detect_book_label(subject_text, category):
    if category == 'korean':
        m = re.search(r'(\d+)\s*(?:китеп|книг)', subject_text, re.I)
        num = int(m.group(1)) if m else 1
        return f'Книга {num}', num
    m = CEFR_RE.search(subject_text)
    if m:
        code = m.group(1).upper()
        code = (CEFR_NORMALIZE.get(code[0], code[0])) + code[1]
        rank = {'A1': 1, 'A2': 2, 'B1': 3, 'B2': 4, 'C1': 5, 'C2': 6}.get(code, 1)
        return code, rank
    # "Общий курс" — используем ранг вне диапазона CEFR (1-6), чтобы не
    # столкнуться по слагу category-book-N с уже созданной книгой уровня.
    return 'Общий курс', 50


def extract_teacher_name(subject_text):
    first = subject_text.split()[0].strip(',.') if subject_text.split() else ''
    return first.capitalize()


def normalize_phone(raw):
    if raw is None:
        return ''
    digits = ''.join(ch for ch in str(raw) if ch.isdigit())
    if not digits:
        return ''
    if len(digits) == 9:
        return '+996' + digits
    if len(digits) == 10 and digits.startswith('0'):
        return '+996' + digits[1:]
    if len(digits) >= 11:
        return '+' + digits[-11:] if not str(raw).strip().startswith('+') else '+' + digits
    return digits  # слишком короткий/битый номер — сохраняем как есть


def parse_side(ws, room_num, fio_col):
    """Разбирает одну колоночную половину листа (fio_col = 1 левая, 5 правая)."""
    blocks = []  # список dict(time, subject, teacher, category, book_label, book_rank, students=[])
    current = None
    expect_subject = False

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    for row in rows:
        val = row[fio_col]
        sval = str(val).strip() if val is not None else ''

        if TIME_RE.match(sval):
            current = None
            expect_subject = True
            pending_time = sval
            continue

        if not sval:
            continue

        if sval == 'ФИО' or sval == '№':
            continue

        if expect_subject:
            subject_text = sval
            expect_subject = False
            if subject_text.lower().startswith('преподовател') or subject_text.lower().startswith('преподавател'):
                current = None  # пустой слот — пропускаем
                continue
            teacher_name = extract_teacher_name(subject_text)
            teacher_key = teacher_name.lower()
            category = detect_category(subject_text, teacher_key)
            book_label, book_rank = detect_book_label(subject_text, category)
            current = {
                'room': room_num,
                'time': pending_time,
                'subject_raw': subject_text,
                'teacher': teacher_name,
                'category': category,
                'book_label': book_label,
                'book_rank': book_rank,
                'students': [],
            }
            blocks.append(current)
            continue

        # строка ученика
        if current is not None:
            phone_raw = row[fio_col + 1] if fio_col + 1 < len(row) else None
            current['students'].append({
                'fio': sval,
                'phone': normalize_phone(phone_raw),
            })

    return blocks


def parse_workbook():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    all_blocks = []
    for ws in wb.worksheets:
        m = DIGIT_RE.search(ws.title)
        room_num = int(m.group(0)) if m else 0
        all_blocks += parse_side(ws, room_num, fio_col=1)
        all_blocks += parse_side(ws, room_num, fio_col=5)
    return all_blocks


def print_summary(blocks):
    total_students = sum(len(b['students']) for b in blocks)
    print(f'Групп (блоков с преподавателем): {len(blocks)}')
    print(f'Учеников всего: {total_students}')
    print()
    for b in blocks:
        print(f"Ауд.{b['room']} {b['time']} | {b['category']:8s} | {b['teacher']:10s} | "
              f"{b['book_label']:10s} | учеников: {len(b['students'])}  (raw: {b['subject_raw']!r})")
        for s in b['students']:
            print(f"     - {s['fio']:30s} {s['phone']}")


# ---------------------------------------------------------------------------
# COMMIT: реальная запись в БД
# ---------------------------------------------------------------------------
def _translit(text):
    mapping = {
        'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
        'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
        'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
        'ф': 'f', 'х': 'kh', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
        'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
    }
    result = []
    for ch in text.lower():
        result.append(mapping.get(ch, ch if ch.isalnum() else ''))
    return ''.join(result)


def commit_import(blocks):
    from django.contrib.auth.models import User
    from django.db import transaction
    from courses.models import Category, Course
    from teachers.models import Teacher
    from lms.models import Group, Book, StudentProfile, UserProfile
    from applications.models import Branch, GroupSet

    created = {
        'categories': 0, 'courses': 0, 'books': 0, 'teachers': 0,
        'groups': 0, 'students': 0, 'skipped_students': 0, 'group_sets': 0,
    }
    credentials = []  # (fio, username, password, group_name)

    with transaction.atomic():
        branch, _ = Branch.objects.get_or_create(
            name='Ош — главный офис', defaults={'address': 'А. Масалиева 44, ТЦ Корона, 3 этаж', 'is_active': True}
        )

        # --- категории и курсы ---
        categories = {}
        courses = {}
        for i, (code, (name_ru, flag, slug)) in enumerate(CATEGORY_META.items()):
            cat, was_created = Category.objects.get_or_create(
                language_code=code,
                defaults={'name': name_ru, 'slug': slug, 'flag_emoji': flag, 'order': i}
            )
            categories[code] = cat
            created['categories'] += was_created

            course_name = {
                'korean': 'Корейский язык с нуля + Подготовка к TOPIK I',
                'english': 'Английский язык & IELTS',
                'german': 'Немецкий язык A1-B1 для Ausbildung в Германии',
                'chinese': 'Китайский язык & HSK',
            }[code]
            course, was_created = Course.objects.get_or_create(
                category=cat, name=course_name,
                defaults={'level': 'beginner', 'duration': '2 месяца', 'lessons_per_week': 3,
                          'price': 2500, 'is_featured': True, 'is_active': True}
            )
            courses[code] = course
            created['courses'] += was_created

        # --- преподаватели (только те, что реально встречаются в Excel) ---
        teachers = {}
        teacher_langs = {}
        for b in blocks:
            teacher_langs.setdefault(b['teacher'], set()).add(CATEGORY_META[b['category']][0])

        for i, (name, langs) in enumerate(sorted(teacher_langs.items())):
            t, was_created = Teacher.objects.get_or_create(
                name=name,
                defaults={'position': 'Преподаватель', 'languages': ', '.join(sorted(langs)),
                          'is_active': True, 'order': i, 'experience_years': 2}
            )
            if not was_created:
                t.languages = ', '.join(sorted(langs))
                t.save(update_fields=['languages'])
            teachers[name] = t
            created['teachers'] += was_created

        # --- книги/уровни ---
        books = {}
        for b in blocks:
            key = (b['category'], b['book_label'])
            if key not in books:
                book, was_created = Book.objects.get_or_create(
                    category=categories[b['category']], name=b['book_label'],
                    defaults={'order': b['book_rank'], 'duration_months': 2, 'price_per_month': 2500, 'is_active': True}
                )
                books[key] = book
                created['books'] += was_created

        # --- группы + ученики ---
        for b in blocks:
            if not b['students']:
                continue
            cat = categories[b['category']]
            course = courses[b['category']]
            teacher = teachers[b['teacher']]
            book = books[(b['category'], b['book_label'])]
            start_time = parse_time(b['time'])
            end_hour = (start_time.hour + 1) % 24
            end_time = dtime(end_hour, start_time.minute)

            group_name = f"{CATEGORY_META[b['category']][0]} — {b['book_label']} — {b['teacher']} — Ауд.{b['room']} {b['time'].replace('-', ':')}"
            group, was_created = Group.objects.get_or_create(
                name=group_name,
                defaults={
                    'course': course, 'teacher': teacher, 'book': book,
                    'start_time': start_time, 'end_time': end_time,
                    'room': f"Ауд. {b['room']}", 'status': 'active', 'branch': branch,
                }
            )
            created['groups'] += was_created

            for s in b['students']:
                parts = s['fio'].split()
                if len(parts) >= 2:
                    last_name, first_name = parts[0], ' '.join(parts[1:])
                else:
                    last_name, first_name = '', parts[0] if parts else 'Ученик'

                if not first_name.strip():
                    created['skipped_students'] += 1
                    continue

                base = (_translit(first_name) + '.' + _translit(last_name)).strip('.')
                if not base:
                    base = ''.join(ch for ch in s['phone'] if ch.isdigit()) or 'student'
                username = base
                n = 1
                while User.objects.filter(username=username).exists():
                    username = f'{base}{n}'
                    n += 1
                password = ''.join(random.choice(string.ascii_lowercase + string.digits) for _ in range(8))

                user = User.objects.create_user(
                    username=username, password=password,
                    first_name=first_name.strip(), last_name=last_name.strip(),
                )
                profile, _ = UserProfile.objects.get_or_create(user=user)
                profile.role = 'student'
                profile.phone = s['phone']
                profile.save(update_fields=['role', 'phone'])
                StudentProfile.objects.create(user=user, group=group, book=book, phone=s['phone'])
                created['students'] += 1
                credentials.append((s['fio'], username, password, group_name))

        # --- открытые наборы для главной страницы (по одному на язык, если есть группы) ---
        seen_langs = set()
        for b in blocks:
            if not b['students'] or b['category'] in seen_langs:
                continue
            seen_langs.add(b['category'])
            gs, was_created = GroupSet.objects.get_or_create(
                course=courses[b['category']], teacher=teachers[b['teacher']], status='open',
                defaults={
                    'days': 'mon,wed,fri',
                    'start_time': parse_time(b['time']),
                    'branch': branch, 'capacity': 12,
                }
            )
            created['group_sets'] += was_created

    return created, credentials


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--commit', action='store_true')
    parser.add_argument('--dry-run', action='store_true')
    args = parser.parse_args()

    blocks = parse_workbook()
    print_summary(blocks)

    if args.commit:
        print()
        print('=== ЗАПИСЬ В БАЗУ ===')
        created, credentials = commit_import(blocks)
        print(created)
        cred_path = os.path.join(BASE, '_student_credentials.csv')
        with open(cred_path, 'w', encoding='utf-8') as f:
            f.write('ФИО;логин;пароль;группа\n')
            for fio, u, p, g in credentials:
                f.write(f'{fio};{u};{p};{g}\n')
        print(f'Логины/пароли сохранены в {cred_path}')
    else:
        print()
        print('(режим предпросмотра — в базу ничего не записано; запустите с --commit для реального импорта)')


if __name__ == '__main__':
    main()
